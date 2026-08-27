
"""MarketLens Export Engine - Excel and PDF report generation.

Exports all data fields produced by the analysis pipeline, including
profitability, AI analysis, consistency sub-scores, forecasting,
marketing, seller info, supplier details, hidden gems, and portfolio
summary.
"""

import json
import logging
from collections import defaultdict
from datetime import datetime
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


def _safe_str(val: Any, max_len: int = 0) -> str:
    if val is None:
        return ""
    s = str(val)
    return s[:max_len] if max_len else s


def _fmt_currency(val: Any) -> str:
    try:
        return f"£{float(val):,.2f}"
    except (TypeError, ValueError):
        return "£0.00"


def _fmt_pct(val: Any) -> str:
    try:
        return f"{float(val):.1f}%"
    except (TypeError, ValueError):
        return "0.0%"


def _list_to_str(val: Any) -> str:
    if isinstance(val, (list, tuple)):
        return ", ".join(str(v) for v in val)
    if val is None:
        return ""
    return str(val)
class ExcelExporter:
    """Export data to a multi-sheet Excel workbook with full field coverage."""

    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    GREEN_FILL  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    RED_FILL    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    GOLD_FILL   = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    BLUE_FILL   = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    THIN_BORDER = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def export_products(
        self,
        products: List[Dict],
        filepath: str,
        hidden_gems: Optional[List[Dict]] = None,
        portfolio_summary: Optional[Dict] = None,
        categories: Optional[List[str]] = None,
        keywords: Optional[List[str]] = None,
    ) -> str:
        wb = Workbook()
        self._create_products_sheet(wb, products)
        self._create_profitability_sheet(wb, products)
        self._create_ai_analysis_sheet(wb, products)
        self._create_consistency_sheet(wb, products)
        self._create_forecasting_sheet(wb, products)
        self._create_marketing_sheet(wb, products)
        self._create_seller_info_sheet(wb, products)
        self._create_supplier_details_sheet(wb, products)
        if hidden_gems:
            self._create_hidden_gems_sheet(wb, hidden_gems)
        if portfolio_summary:
            self._create_portfolio_summary_sheet(wb, portfolio_summary)
        self._create_summary_sheet(wb, products, hidden_gems, portfolio_summary)
        wb.save(filepath)
        logger.info("Excel export saved to %s (%d products)", filepath, len(products))
        return filepath

    def _sorted(self, products: List[Dict]) -> List[Dict]:
        return sorted(products, key=lambda x: x.get("ai_score", 0), reverse=True)

    def _write_header(self, ws, headers: List[str], fill=None, font=None):
        fill = fill or self.HEADER_FILL
        font = font or self.HEADER_FONT
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.THIN_BORDER

    def _auto_width(self, ws, widths: Dict[int, int]):
        for col_idx, w in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = w
    def _create_products_sheet(self, wb, products):
        ws = wb.active
        ws.title = "Products"
        headers = [
            "Rank", "Product Name", "ASIN", "Brand", "Category", "Price", "Rating",
            "Reviews", "Source", "URL", "Image",
            "Margin %", "Estimated Profit", "Tier", "Viable",
            "AI Score", "AI Recommendation", "Suggested Action",
            "Consistency", "Consistency Tier", "Traffic Light",
            "Demand Pattern", "Portfolio Type",
            "CAGR %", "Outlook", "Evergreen Prob",
            "Priority Rank", "Priority Tier", "Priority Action",
            "Validated", "Gated",
            "Supplier Name", "Supplier Price",
        ]
        self._write_header(ws, headers)
        for i, p in enumerate(self._sorted(products), 1):
            row = i + 1
            pri = p.get("priority", {}) if isinstance(p.get("priority"), dict) else {}
            fc = p.get("forecast", {}) if isinstance(p.get("forecast"), dict) else {}
            vals = [
                i, _safe_str(p.get("name", p.get("title", "")), 80),
                p.get("asin", ""), p.get("brand_name", ""),
                p.get("category", ""),
                p.get("amazon_price", p.get("price", 0)), p.get("rating", 0),
                p.get("review_count", 0), p.get("source", ""),
                p.get("url", ""), p.get("image", ""),
                p.get("estimated_margin_pct", 0), p.get("estimated_profit", 0),
                p.get("tier", ""), "Yes" if p.get("viable") else "No",
                p.get("ai_score", 0),
                _safe_str(p.get("ai_recommendation", ""), 80),
                p.get("suggested_action", ""),
                p.get("consistency_score", 0), p.get("consistency_tier", ""),
                p.get("traffic_light", "RED"), p.get("demand_pattern", ""),
                p.get("portfolio_type", ""),
                fc.get("cagr_pct", 0), fc.get("overall_outlook", ""),
                fc.get("evergreen_probability", 0),
                pri.get("rank", p.get("priority_rank", "")),
                pri.get("tier", ""), pri.get("action", ""),
                "Yes" if p.get("validated") else "No",
                "Yes" if p.get("is_gated") else "No",
                p.get("supplier_name", ""), p.get("supplier_price", 0),
            ]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = self.THIN_BORDER
                cell.alignment = Alignment(
                    horizontal="center" if col not in (2, 16) else "left",
                    wrap_text=True,
                )
                if col == 12:
                    cell.number_format = '0.0%'
                    cell.value = (val / 100) if val else 0
                    v = val if val else 0
                    if v >= 40: cell.fill = self.GREEN_FILL
                    elif v >= 25: cell.fill = self.YELLOW_FILL
                    else: cell.fill = self.RED_FILL
                elif col in (13, 32):
                    cell.number_format = '£#,##0.00'
                elif col in (16, 26):
                    cell.number_format = '0.0%'
                elif col == 21:
                    if val == "GREEN": cell.fill = self.GREEN_FILL
                    elif val == "YELLOW": cell.fill = self.YELLOW_FILL
                    elif val == "RED": cell.fill = self.RED_FILL
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"
        self._auto_width(ws, {
            1: 6, 2: 50, 3: 14, 4: 16, 5: 14, 6: 10, 7: 8, 8: 10,
            9: 10, 10: 50, 11: 40, 12: 10, 13: 12, 14: 10, 15: 8,
            16: 10, 17: 40, 18: 14, 19: 12, 20: 14, 21: 12,
            22: 14, 23: 12, 24: 10, 25: 14, 26: 12, 27: 12,
            28: 10, 29: 12, 30: 10, 31: 20, 32: 12,
        })
    def _create_profitability_sheet(self, wb, products):
        ws = wb.create_sheet("Profitability")
        headers = [
            "ASIN", "Product Name", "Category", "Amazon Price",
            "Estimated Supplier Cost", "FBA Fees", "Total Cost",
            "Estimated Profit", "Margin %", "Viable", "Tier",
        ]
        fill = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
        self._write_header(ws, headers, fill=fill)
        for i, p in enumerate(self._sorted(products), 2):
            vals = [
                p.get("asin", ""), _safe_str(p.get("name", p.get("title", "")), 60),
                p.get("category", ""), p.get("amazon_price", p.get("price", 0)),
                p.get("estimated_supplier_cost", 0), p.get("fba_fees", 0),
                p.get("total_cost", 0), p.get("estimated_profit", 0),
                p.get("estimated_margin_pct", 0) / 100,
                "Yes" if p.get("viable") else "No", p.get("tier", ""),
            ]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=i, column=col, value=val)
                cell.border = self.THIN_BORDER
                if col in (4, 5, 6, 7, 8):
                    cell.number_format = '£#,##0.00'
                elif col == 9:
                    cell.number_format = '0.0%'
                    v = (val * 100) if val else 0
                    if v >= 40: cell.fill = self.GREEN_FILL
                    elif v >= 25: cell.fill = self.YELLOW_FILL
                    else: cell.fill = self.RED_FILL
                elif col == 10:
                    cell.fill = self.GREEN_FILL if val == "Yes" else self.RED_FILL
        self._auto_width(ws, {1: 14, 2: 50, 3: 14, 4: 12, 5: 18, 6: 12, 7: 12, 8: 14, 9: 10, 10: 8, 11: 10})
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

    def _create_ai_analysis_sheet(self, wb, products):
        ws = wb.create_sheet("AI Analysis")
        headers = [
            "ASIN", "Product Name", "AI Score", "AI Recommendation",
            "AI Confidence", "Market Demand", "Competition Level",
            "Suggested Action", "Price Analysis", "Niche Opportunity",
            "Improvement Areas", "Risk Factors", "Est. Monthly Revenue",
        ]
        fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
        self._write_header(ws, headers, fill=fill)
        for i, p in enumerate(self._sorted(products), 2):
            vals = [
                p.get("asin", ""), _safe_str(p.get("name", p.get("title", "")), 60),
                p.get("ai_score", 0), _safe_str(p.get("ai_recommendation", ""), 120),
                p.get("ai_confidence", 0), p.get("market_demand", ""),
                p.get("competition_level", ""), p.get("suggested_action", ""),
                _safe_str(p.get("price_analysis", ""), 200),
                p.get("niche_opportunity", ""),
                _list_to_str(p.get("improvement_areas", [])),
                _list_to_str(p.get("risk_factors", [])),
                _safe_str(p.get("estimated_monthly_revenue", ""), 40),
            ]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=i, column=col, value=val)
                cell.border = self.THIN_BORDER
                cell.alignment = Alignment(wrap_text=True)
                if col in (3, 5): cell.number_format = '0.0%'
        self._auto_width(ws, {1: 14, 2: 50, 3: 10, 4: 40, 5: 12, 6: 14, 7: 14, 8: 16, 9: 50, 10: 14, 11: 40, 12: 40, 13: 18})
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"
    def _create_consistency_sheet(self, wb, products):
        ws = wb.create_sheet("Consistency")
        headers = [
            "ASIN", "Product Name", "Consistency Score", "Consistency Tier",
            "Traffic Light", "Demand Pattern", "Portfolio Type",
            "Demand Stability", "Price Stability", "Review Growth Score",
            "Margin Consistency", "Seasonal Predictability",
            "Market Longevity", "Competitive Moat",
            "Long-term Score", "Short-term Score", "Macro Trends",
        ]
        fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        self._write_header(ws, headers, fill=fill)
        for i, p in enumerate(self._sorted(products), 2):
            vals = [
                p.get("asin", ""), _safe_str(p.get("name", p.get("title", "")), 60),
                p.get("consistency_score", 0), p.get("consistency_tier", ""),
                p.get("traffic_light", "RED"), p.get("demand_pattern", ""),
                p.get("portfolio_type", ""),
                p.get("demand_stability", 0), p.get("price_stability", 0),
                p.get("review_growth_score", 0), p.get("margin_consistency", 0),
                p.get("seasonal_predictability", 0), p.get("market_longevity", 0),
                p.get("competitive_moat", 0), p.get("long_term_score", 0),
                p.get("short_term_score", 0), _list_to_str(p.get("macro_trends", [])),
            ]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=i, column=col, value=val)
                cell.border = self.THIN_BORDER
                cell.alignment = Alignment(wrap_text=True)
                if col in (3, 8, 9, 10, 11, 12, 13, 14, 15, 16):
                    cell.number_format = '0.0%'
                elif col == 5:
                    if val == "GREEN": cell.fill = self.GREEN_FILL
                    elif val == "YELLOW": cell.fill = self.YELLOW_FILL
                    elif val == "RED": cell.fill = self.RED_FILL
        self._auto_width(ws, {1: 14, 2: 50, 3: 14, 4: 14, 5: 12, 6: 14, 7: 12, 8: 14, 9: 14, 10: 14, 11: 14, 12: 16, 13: 14, 14: 14, 15: 14, 16: 14, 17: 30})
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

    def _create_forecasting_sheet(self, wb, products):
        ws = wb.create_sheet("Forecasting")
        headers = [
            "ASIN", "Product Name", "Current Monthly Sales",
            "CAGR %", "CAGR (raw)", "Overall Outlook",
            "Peak Month", "Trough Month", "Evergreen Probability",
            "Macro Growth Impact", "Trend Slope", "Seasonal Strength",
            "Year 1 Total", "Year 2 Total", "Year 3 Total",
            "Year 4 Total", "Year 5 Total",
        ]
        fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
        self._write_header(ws, headers, fill=fill)
        for i, p in enumerate(self._sorted(products), 2):
            fc = p.get("forecast", {})
            if not isinstance(fc, dict): fc = {}
            yearly = fc.get("yearly_forecast", [])
            if not isinstance(yearly, list): yearly = []
            vals = [
                p.get("asin", ""), _safe_str(p.get("name", p.get("title", "")), 60),
                fc.get("current_monthly_sales", 0), fc.get("cagr_pct", 0),
                fc.get("cagr", 0), fc.get("overall_outlook", ""),
                fc.get("peak_month", ""), fc.get("trough_month", ""),
                fc.get("evergreen_probability", 0), fc.get("macro_growth_impact", 0),
                fc.get("trend_slope", 0), fc.get("seasonal_strength", 0),
                yearly[0].get("yearly_total", 0) if len(yearly) > 0 else 0,
                yearly[1].get("yearly_total", 0) if len(yearly) > 1 else 0,
                yearly[2].get("yearly_total", 0) if len(yearly) > 2 else 0,
                yearly[3].get("yearly_total", 0) if len(yearly) > 3 else 0,
                yearly[4].get("yearly_total", 0) if len(yearly) > 4 else 0,
            ]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=i, column=col, value=val)
                cell.border = self.THIN_BORDER
                cell.alignment = Alignment(wrap_text=True)
                if col == 4: cell.number_format = '0.00%'
                elif col in (5, 10, 11, 12): cell.number_format = '0.000'
                elif col == 9: cell.number_format = '0%'
        self._auto_width(ws, {1: 14, 2: 50, 3: 16, 4: 10, 5: 10, 6: 16, 7: 12, 8: 12, 9: 16, 10: 16, 11: 12, 12: 14, 13: 14, 14: 14, 15: 14, 16: 14, 17: 14})
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"
    def _create_marketing_sheet(self, wb, products):
        ws = wb.create_sheet("Marketing")
        headers = [
            "ASIN", "Product Name", "Marketing Score", "Summary",
            "Problem Count", "High Severity", "Medium Severity",
            "Solution Count", "Strategy Count", "Top Strategies",
            "Problem Details", "Solution Details",
        ]
        fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        self._write_header(ws, headers, fill=fill)
        for i, p in enumerate(self._sorted(products), 2):
            mkt = p.get("marketing", {})
            if not isinstance(mkt, dict): mkt = {}
            problems = mkt.get("problems", [])
            solutions = mkt.get("solutions", [])
            strategies = mkt.get("recommended_strategies", [])
            high_sev = sum(1 for pr in problems if isinstance(pr, dict) and pr.get("severity") == "high")
            med_sev = sum(1 for pr in problems if isinstance(pr, dict) and pr.get("severity") == "medium")
            strat_names = [s.get("name", "") for s in strategies if isinstance(s, dict)]
            prob_str = "; ".join(f"{pr.get('problem', '')} [{pr.get('severity', '')}]" for pr in problems if isinstance(pr, dict))
            sol_str = "; ".join(f"{s.get('solution', '')} ({s.get('priority', '')})" for s in solutions if isinstance(s, dict))
            vals = [
                p.get("asin", ""), _safe_str(p.get("name", p.get("title", "")), 60),
                mkt.get("marketing_score", 0), _safe_str(mkt.get("summary", ""), 200),
                len(problems), high_sev, med_sev, len(solutions), len(strategies),
                ", ".join(strat_names[:5]),
                _safe_str(prob_str, 500), _safe_str(sol_str, 500),
            ]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=i, column=col, value=val)
                cell.border = self.THIN_BORDER
                cell.alignment = Alignment(wrap_text=True)
                if col == 3: cell.number_format = '0.0%'
        self._auto_width(ws, {1: 14, 2: 50, 3: 14, 4: 60, 5: 12, 6: 12, 7: 14, 8: 12, 9: 12, 10: 40, 11: 60, 12: 60})
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"
    def _create_seller_info_sheet(self, wb, products):
        ws = wb.create_sheet("Seller Info")
        headers = [
            "ASIN", "Product Name",
            "Seller Name", "Seller Rating", "Seller Reviews",
            "Fulfillment", "Brand", "Manufacturer",
            "BSR", "Monthly Sales Est", "Competition Level",
            "Prime", "Amazon Retail", "Location",
            "Product Weight", "Dimensions",
            "Date First Available", "Amazon Choice",
            "Buy Box Winner", "Num Sellers",
            "Has Coupon", "Coupon Value",
            "Return Policy", "Warranty", "Stock Status",
        ]
        fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        self._write_header(ws, headers, fill=fill)
        for i, p in enumerate(self._sorted(products), 2):
            s = p.get("seller_info", {})
            if not isinstance(s, dict): s = {}
            vals = [
                p.get("asin", ""), _safe_str(p.get("name", p.get("title", "")), 60),
                s.get("seller_name", "N/A"), s.get("seller_rating", 0),
                s.get("seller_reviews", 0),
                "FBA" if s.get("is_fba") else "FBM",
                s.get("brand", "N/A"), s.get("manufacturer", "N/A"),
                s.get("bsr", 0), s.get("monthly_sales_est", 0),
                s.get("competition_level", "N/A"),
                "Yes" if s.get("is_prime") else "No",
                "Yes" if s.get("is_amazon_retail") else "No",
                s.get("seller_location", "N/A"),
                s.get("product_weight", "N/A"), s.get("dimensions", "N/A"),
                s.get("date_first_available", "N/A"),
                "Yes" if s.get("amazon_choice") else "No",
                _safe_str(s.get("buy_box_winner", ""), 60),
                s.get("num_sellers", 0),
                "Yes" if s.get("has_coupon") else "No", s.get("coupon_value", "N/A"),
                _safe_str(s.get("return_policy", ""), 80),
                _safe_str(s.get("warranty", ""), 80), s.get("stock_status", "N/A"),
            ]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=i, column=col, value=val)
                cell.border = self.THIN_BORDER
                cell.alignment = Alignment(wrap_text=True)
                if col == 6:
                    cell.fill = self.GREEN_FILL if val == "FBA" else self.YELLOW_FILL
        self._auto_width(ws, {c: 15 for c in range(1, 26)})
        ws.column_dimensions['B'].width = 50
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"
    def _create_supplier_details_sheet(self, wb, products):
        ws = wb.create_sheet("Supplier Details")
        headers = [
            "ASIN", "Product Name", "Category",
            "Supplier Name", "Supplier Company",
            "Supplier Price", "Supplier MOQ", "Supplier Lead Time",
            "Supplier Payment", "Supplier Rating",
            "Supplier Email", "Supplier Phone",
            "Supplier WhatsApp", "Supplier Website",
            "Supplier Location", "Supplier Price Source", "Bulk Prices",
        ]
        fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
        self._write_header(ws, headers, fill=fill)
        for i, p in enumerate(self._sorted(products), 2):
            bulk = p.get("bulk_prices", {})
            if isinstance(bulk, dict):
                bulk_str = "; ".join(f"{k}: £{v}" for k, v in bulk.items())
            else:
                bulk_str = str(bulk) if bulk else ""
            vals = [
                p.get("asin", ""), _safe_str(p.get("name", p.get("title", "")), 60),
                p.get("category", ""),
                p.get("supplier_name", "N/A"), p.get("supplier_company", "N/A"),
                p.get("supplier_price", 0), p.get("supplier_moq", 0),
                p.get("supplier_lead_time", "N/A"), p.get("supplier_payment", "N/A"),
                p.get("supplier_rating", 0),
                p.get("supplier_email", ""), p.get("supplier_phone", ""),
                p.get("supplier_whatsapp", ""), p.get("supplier_website", ""),
                p.get("supplier_location", "N/A"),
                p.get("supplier_price_source", "N/A"), _safe_str(bulk_str, 200),
            ]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=i, column=col, value=val)
                cell.border = self.THIN_BORDER
                cell.alignment = Alignment(wrap_text=True)
                if col == 6: cell.number_format = '£#,##0.00'
        self._auto_width(ws, {c: 16 for c in range(1, 18)})
        ws.column_dimensions['B'].width = 50
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

    def _create_hidden_gems_sheet(self, wb, gems):
        ws = wb.create_sheet("Hidden Gems")
        headers = [
            "Rank", "Name", "ASIN", "Price", "Rating", "Reviews",
            "Category", "Potential Score", "Low Competition",
            "Trend Score", "Social Score", "Niche Score",
            "Margin Score", "Type", "Reasons", "URL", "Image",
        ]
        fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        font = Font(name="Calibri", bold=True, color="000000", size=11)
        self._write_header(ws, headers, fill=fill, font=font)
        for i, gem in enumerate(sorted(gems, key=lambda x: x.get("potential_score", 0), reverse=True), 1):
            row = i + 1
            vals = [
                i, _safe_str(gem.get("name", ""), 80), gem.get("asin", ""),
                gem.get("amazon_price", 0), gem.get("rating", 0),
                gem.get("review_count", 0), gem.get("category", ""),
                gem.get("potential_score", 0), gem.get("low_competition", 0),
                gem.get("trend_score", 0), gem.get("social_score", 0),
                gem.get("niche_score", 0), gem.get("margin_score", 0),
                gem.get("type", ""), _list_to_str(gem.get("reasons", [])),
                gem.get("url", ""), gem.get("image", ""),
            ]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = self.THIN_BORDER
                cell.alignment = Alignment(wrap_text=True)
                if col == 4: cell.number_format = '£#,##0.00'
                elif col in (8, 9, 10, 11, 12, 13): cell.number_format = '0.0%'
        self._auto_width(ws, {1: 6, 2: 50, 3: 14, 4: 10, 5: 8, 6: 10, 7: 14, 8: 14, 9: 14, 10: 12, 11: 12, 12: 12, 13: 12, 14: 14, 15: 60, 16: 50, 17: 40})
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

    def _create_portfolio_summary_sheet(self, wb, portfolio):
        ws = wb.create_sheet("Portfolio Summary")
        ws.cell(row=1, column=1, value="PORTFOLIO SUMMARY").font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        row = 4
        for key, val in portfolio.items():
            ws.cell(row=row, column=1, value=str(key)).font = Font(bold=True)
            if isinstance(val, (dict, list)):
                ws.cell(row=row, column=2, value=json.dumps(val, default=str)[:200])
            else:
                ws.cell(row=row, column=2, value=str(val))
            row += 1
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 80
    def _create_summary_sheet(self, wb, products, hidden_gems=None, portfolio_summary=None):
        ws = wb.create_sheet("Summary")
        ws.cell(row=1, column=1, value="MARKETLENS ANALYSIS SUMMARY").font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        total = len(products)
        ws.cell(row=4, column=1, value="PRODUCT OVERVIEW").font = Font(bold=True, size=12)
        overview = [
            ("Total Products", total),
            ("Products with AI Score", len([p for p in products if p.get("ai_score", 0) > 0])),
            ("High Margin (40%+)", len([p for p in products if p.get("estimated_margin_pct", 0) >= 40])),
            ("Medium Margin (25-40%)", len([p for p in products if 25 <= p.get("estimated_margin_pct", 0) < 40])),
            ("Low Margin (<25%)", len([p for p in products if 0 < p.get("estimated_margin_pct", 0) < 25])),
            ("Viable Products", len([p for p in products if p.get("viable")])),
            ("Validated Products", len([p for p in products if p.get("validated")])),
        ]
        for i, (label, val) in enumerate(overview, 5):
            ws.cell(row=i, column=1, value=label).font = Font(bold=True)
            ws.cell(row=i, column=2, value=val)
        ws.cell(row=13, column=1, value="TRAFFIC LIGHT DISTRIBUTION").font = Font(bold=True, size=12)
        for i, (label, colour) in enumerate([("Evergreen (GREEN)", "GREEN"), ("Seasonal (YELLOW)", "YELLOW"), ("Volatile (RED)", "RED")], 14):
            count = len([p for p in products if p.get("traffic_light") == colour])
            ws.cell(row=i, column=1, value=label).font = Font(bold=True)
            ws.cell(row=i, column=2, value=count)
        ws.cell(row=18, column=1, value="PRIORITY DISTRIBUTION").font = Font(bold=True, size=12)
        for i, tier in enumerate(["CRITICAL", "HIGH", "MEDIUM", "LOW", "MINIMAL"], 19):
            count = len([p for p in products if isinstance(p.get("priority"), dict) and p["priority"].get("tier") == tier])
            ws.cell(row=i, column=1, value=tier).font = Font(bold=True)
            ws.cell(row=i, column=2, value=count)
        ws.cell(row=25, column=1, value="PORTFOLIO TYPE DISTRIBUTION").font = Font(bold=True, size=12)
        for i, ptype in enumerate(["ANCHOR", "GROWTH", "BALANCED", "WATCHLIST"], 26):
            count = len([p for p in products if p.get("portfolio_type") == ptype])
            ws.cell(row=i, column=1, value=ptype).font = Font(bold=True)
            ws.cell(row=i, column=2, value=count)
        ws.cell(row=31, column=1, value="CONSISTENCY TIER DISTRIBUTION").font = Font(bold=True, size=12)
        for i, tier in enumerate(["platinum", "gold", "silver", "bronze", "unrated"], 32):
            count = len([p for p in products if p.get("consistency_tier", "").lower() == tier])
            ws.cell(row=i, column=1, value=tier.title()).font = Font(bold=True)
            ws.cell(row=i, column=2, value=count)
        ws.cell(row=38, column=1, value="KEY AVERAGES").font = Font(bold=True, size=12)
        avgs = [
            ("Avg Margin", "{:.1f}%".format(sum(p.get("estimated_margin_pct", 0) for p in products) / max(total, 1))),
            ("Avg AI Score", "{:.1%}".format(sum(p.get("ai_score", 0) for p in products) / max(total, 1))),
            ("Avg Consistency", "{:.1%}".format(sum(p.get("consistency_score", 0) for p in products) / max(total, 1))),
        ]
        for i, (label, val) in enumerate(avgs, 39):
            ws.cell(row=i, column=1, value=label).font = Font(bold=True)
            ws.cell(row=i, column=2, value=val)
        if hidden_gems:
            ws.cell(row=45, column=1, value="HIDDEN GEMS").font = Font(bold=True, size=12)
            ws.cell(row=46, column=1, value="Total Hidden Gems").font = Font(bold=True)
            ws.cell(row=46, column=2, value=len(hidden_gems))
            ws.cell(row=47, column=1, value="Emerging Niches").font = Font(bold=True)
            ws.cell(row=47, column=2, value=len([g for g in hidden_gems if g.get("type") == "emerging_niche"]))
            ws.cell(row=48, column=1, value="Hidden Gem Products").font = Font(bold=True)
            ws.cell(row=48, column=2, value=len([g for g in hidden_gems if g.get("type") == "hidden_gem"]))
        ws.cell(row=50, column=1, value="TOP 10 PRODUCTS").font = Font(bold=True, size=12)
        top = self._sorted(products)[:10]
        for i, p in enumerate(top, 51):
            ws.cell(row=i, column=1, value=f"#{i - 50}")
            ws.cell(row=i, column=2, value=_safe_str(p.get("name", p.get("title", "")), 50))
            ws.cell(row=i, column=3, value="AI: {:.0%}".format(p.get("ai_score", 0)))
            ws.cell(row=i, column=4, value="Margin: {:.0f}%".format(p.get("estimated_margin_pct", 0)))
            ws.cell(row=i, column=5, value=p.get("traffic_light", ""))
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 55
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
class PDFExporter:
    """Export comprehensive reports to PDF."""

    def __init__(self):
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import inch
        from reportlab.platypus import (
            PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
        )
        self.colors = colors
        self.letter = letter
        self.inch = inch
        self._ParagraphStyle = ParagraphStyle
        self._Paragraph = Paragraph
        self._SimpleDocTemplate = SimpleDocTemplate
        self._Spacer = Spacer
        self._Table = Table
        self._TableStyle = TableStyle
        self._PageBreak = PageBreak
        self.styles = getSampleStyleSheet()
        self._add_custom_styles()

    def _add_custom_styles(self):
        self.styles.add(self._ParagraphStyle(name='TitleMain', parent=self.styles['Title'], fontSize=22, spaceAfter=30, textColor=self.colors.HexColor('#1F4E79')))
        self.styles.add(self._ParagraphStyle(name='SectionHeader', parent=self.styles['Heading2'], fontSize=14, spaceAfter=12, textColor=self.colors.HexColor('#2E75B6')))
        self.styles.add(self._ParagraphStyle(name='SubSection', parent=self.styles['Heading3'], fontSize=11, spaceAfter=8, textColor=self.colors.HexColor('#548235')))
        self.styles.add(self._ParagraphStyle(name='SmallText', parent=self.styles['Normal'], fontSize=8, leading=10))

    def export_report(self, products, filepath, title="MarketLens Analysis Report", hidden_gems=None, portfolio_summary=None):
        doc = self._SimpleDocTemplate(filepath, pagesize=self.letter, topMargin=0.75*self.inch, bottomMargin=0.75*self.inch, leftMargin=0.75*self.inch, rightMargin=0.75*self.inch)
        story = []
        story.append(self._Paragraph(title, self.styles['TitleMain']))
        story.append(self._Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", self.styles['Normal']))
        story.append(self._Spacer(1, 20))
        story.extend(self._create_executive_summary(products))
        story.append(self._PageBreak())
        story.extend(self._create_top_products_section(products))
        story.append(self._PageBreak())
        story.extend(self._create_profitability_section(products))
        story.append(self._PageBreak())
        story.extend(self._create_ai_insights_section(products))
        story.append(self._PageBreak())
        story.extend(self._create_consistency_section(products))
        story.append(self._PageBreak())
        story.extend(self._create_forecast_section(products))
        story.append(self._PageBreak())
        story.extend(self._create_marketing_section(products))
        story.append(self._PageBreak())
        story.extend(self._create_supplier_section(products))
        story.append(self._PageBreak())
        story.extend(self._create_category_section(products))
        if hidden_gems:
            story.append(self._PageBreak())
            story.extend(self._create_hidden_gems_section(hidden_gems))
        doc.build(story)
        logger.info("PDF export saved to %s (%d products)", filepath, len(products))
        return filepath

    def _make_table(self, data, col_widths, header_color='#1F4E79'):
        table = self._Table(data, colWidths=col_widths)
        table.setStyle(self._TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), self.colors.HexColor(header_color)),
            ('TEXTCOLOR', (0, 0), (-1, 0), self.colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, self.colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [self.colors.white, self.colors.HexColor('#F8F9FA')]),
        ]))
        return table

    def _create_executive_summary(self, products):
        elements = []
        elements.append(self._Paragraph("Executive Summary", self.styles['SectionHeader']))
        total = len(products)
        high_margin = len([p for p in products if p.get("estimated_margin_pct", 0) >= 40])
        high_ai = len([p for p in products if p.get("ai_score", 0) >= 0.7])
        green = len([p for p in products if p.get("traffic_light") == "GREEN"])
        viable = len([p for p in products if p.get("viable")])
        avg_margin = sum(p.get("estimated_margin_pct", 0) for p in products) / max(total, 1)
        avg_ai = sum(p.get("ai_score", 0) for p in products) / max(total, 1)
        data = [
            ["Metric", "Value"], ["Total Products", str(total)],
            ["High AI Score (70%+)", str(high_ai)], ["High Margin (40%+)", str(high_margin)],
            ["Viable Products", str(viable)], ["Evergreen (GREEN)", str(green)],
            ["Avg Margin", f"{avg_margin:.1f}%"], ["Avg AI Score", f"{avg_ai:.0%}"],
        ]
        table = self._Table(data, colWidths=[3*self.inch, 2*self.inch])
        table.setStyle(self._TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), self.colors.HexColor('#1F4E79')),
            ('TEXTCOLOR', (0, 0), (-1, 0), self.colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), self.colors.HexColor('#F0F4F8')),
            ('GRID', (0, 0), (-1, -1), 1, self.colors.grey),
        ]))
        elements.append(table)
        elements.append(self._Spacer(1, 20))
        return elements
    def _create_top_products_section(self, products):
        elements = []
        elements.append(self._Paragraph("Top 20 Products", self.styles['SectionHeader']))
        top = sorted(products, key=lambda x: x.get("ai_score", 0), reverse=True)[:20]
        data = [["#", "Product", "Category", "Price", "Margin", "AI Score", "Traffic", "Priority"]]
        for i, p in enumerate(top, 1):
            pri = p.get("priority", {}) if isinstance(p.get("priority"), dict) else {}
            data.append([str(i), self._Paragraph(_safe_str(p.get("name", p.get("title", "")), 35), self.styles['SmallText']),
                _safe_str(p.get("category", ""), 10), _fmt_currency(p.get("amazon_price", 0)),
                _fmt_pct(p.get("estimated_margin_pct", 0)), f"{p.get('ai_score', 0):.0%}",
                p.get("traffic_light", "N/A"), pri.get("tier", "")])
        table = self._make_table(data, [0.3*self.inch, 2.2*self.inch, 0.7*self.inch, 0.7*self.inch, 0.6*self.inch, 0.6*self.inch, 0.6*self.inch, 0.7*self.inch])
        elements.append(table)
        return elements

    def _create_profitability_section(self, products):
        elements = []
        elements.append(self._Paragraph("Profitability Analysis", self.styles['SectionHeader']))
        data = [["#", "Product", "Price", "Supplier Cost", "FBA Fees", "Profit", "Margin", "Viable"]]
        top = sorted(products, key=lambda x: x.get("estimated_margin_pct", 0), reverse=True)[:25]
        for i, p in enumerate(top, 1):
            data.append([str(i), self._Paragraph(_safe_str(p.get("name", p.get("title", "")), 35), self.styles['SmallText']),
                _fmt_currency(p.get("amazon_price", 0)), _fmt_currency(p.get("estimated_supplier_cost", 0)),
                _fmt_currency(p.get("fba_fees", 0)), _fmt_currency(p.get("estimated_profit", 0)),
                _fmt_pct(p.get("estimated_margin_pct", 0)), "Yes" if p.get("viable") else "No"])
        table = self._make_table(data, [0.3*self.inch, 2.0*self.inch, 0.7*self.inch, 0.8*self.inch, 0.6*self.inch, 0.7*self.inch, 0.6*self.inch, 0.5*self.inch])
        elements.append(table)
        return elements

    def _create_ai_insights_section(self, products):
        elements = []
        elements.append(self._Paragraph("AI Analysis Insights", self.styles['SectionHeader']))
        data = [["#", "Product", "AI Score", "Confidence", "Market Demand", "Competition", "Action"]]
        top = sorted(products, key=lambda x: x.get("ai_score", 0), reverse=True)[:20]
        for i, p in enumerate(top, 1):
            data.append([str(i), self._Paragraph(_safe_str(p.get("name", p.get("title", "")), 35), self.styles['SmallText']),
                f"{p.get('ai_score', 0):.0%}", f"{p.get('ai_confidence', 0):.0%}",
                p.get("market_demand", "N/A"), p.get("competition_level", "N/A"), p.get("suggested_action", "N/A")])
        table = self._make_table(data, [0.3*self.inch, 2.2*self.inch, 0.6*self.inch, 0.7*self.inch, 0.9*self.inch, 0.9*self.inch, 0.9*self.inch], header_color='#7030A0')
        elements.append(table)
        elements.append(self._Spacer(1, 15))
        elements.append(self._Paragraph("Top 5 - Detailed Insights", self.styles['SubSection']))
        for i, p in enumerate(top[:5], 1):
            name = _safe_str(p.get("name", p.get("title", "")), 60)
            elements.append(self._Paragraph(f"<b>#{i} {name}</b>", self.styles['Normal']))
            elements.append(self._Paragraph(f"Score: {p.get('ai_score', 0):.0%} | Market Demand: {p.get('market_demand', 'N/A')} | Competition: {p.get('competition_level', 'N/A')}", self.styles['SmallText']))
            elements.append(self._Paragraph(f"Recommendation: {p.get('ai_recommendation', 'N/A')}", self.styles['SmallText']))
            areas = p.get("improvement_areas", [])
            if areas:
                elements.append(self._Paragraph(f"Improvement Areas: {', '.join(str(a) for a in areas)}", self.styles['SmallText']))
            elements.append(self._Spacer(1, 8))
        return elements
    def _create_consistency_section(self, products):
        elements = []
        elements.append(self._Paragraph("Consistency & Demand Analysis", self.styles['SectionHeader']))
        data = [["#", "Product", "Consistency", "Tier", "Pattern", "Traffic", "Portfolio"]]
        top = sorted(products, key=lambda x: x.get("consistency_score", 0), reverse=True)[:20]
        for i, p in enumerate(top, 1):
            data.append([str(i), self._Paragraph(_safe_str(p.get("name", p.get("title", "")), 35), self.styles['SmallText']),
                f"{p.get('consistency_score', 0):.0%}", p.get("consistency_tier", "N/A"),
                p.get("demand_pattern", "N/A"), p.get("traffic_light", "N/A"), p.get("portfolio_type", "N/A")])
        table = self._make_table(data, [0.3*self.inch, 2.2*self.inch, 0.9*self.inch, 0.7*self.inch, 0.8*self.inch, 0.7*self.inch, 0.8*self.inch], header_color='#00B050')
        elements.append(table)
        elements.append(self._Spacer(1, 15))
        elements.append(self._Paragraph("Sub-Score Breakdown (Top 5)", self.styles['SubSection']))
        sub_data = [["Product", "Demand Stab", "Price Stab", "Review Gr", "Margin Con", "Season Pr", "Longevity", "Moat"]]
        for p in top[:5]:
            sub_data.append([self._Paragraph(_safe_str(p.get("name", p.get("title", "")), 25), self.styles['SmallText']),
                f"{p.get('demand_stability', 0):.0%}", f"{p.get('price_stability', 0):.0%}",
                f"{p.get('review_growth_score', 0):.0%}", f"{p.get('margin_consistency', 0):.0%}",
                f"{p.get('seasonal_predictability', 0):.0%}", f"{p.get('market_longevity', 0):.0%}",
                f"{p.get('competitive_moat', 0):.0%}"])
        sub_table = self._make_table(sub_data, [1.2*self.inch] + [0.8*self.inch]*7, header_color='#00B050')
        elements.append(sub_table)
        return elements

    def _create_forecast_section(self, products):
        elements = []
        elements.append(self._Paragraph("5-Year Demand Forecast", self.styles['SectionHeader']))
        data = [["#", "Product", "Monthly Sales", "CAGR", "Outlook", "Peak", "Evergreen"]]
        top = sorted(products, key=lambda x: x.get("ai_score", 0), reverse=True)[:20]
        for i, p in enumerate(top, 1):
            fc = p.get("forecast", {})
            if not isinstance(fc, dict): fc = {}
            data.append([str(i), self._Paragraph(_safe_str(p.get("name", p.get("title", "")), 35), self.styles['SmallText']),
                str(fc.get("current_monthly_sales", 0)), f"{fc.get('cagr_pct', 0):.1f}%",
                fc.get("overall_outlook", "N/A"), fc.get("peak_month", "N/A"),
                f"{fc.get('evergreen_probability', 0):.0%}"])
        table = self._make_table(data, [0.3*self.inch, 2.2*self.inch, 0.9*self.inch, 0.6*self.inch, 0.9*self.inch, 0.7*self.inch, 0.8*self.inch], header_color='#0070C0')
        elements.append(table)
        return elements

    def _create_marketing_section(self, products):
        elements = []
        elements.append(self._Paragraph("Marketing Analysis", self.styles['SectionHeader']))
        data = [["#", "Product", "Score", "Problems", "Solutions", "Strategies", "Summary"]]
        top = sorted(products, key=lambda x: x.get("ai_score", 0), reverse=True)[:15]
        for i, p in enumerate(top, 1):
            mkt = p.get("marketing", {})
            if not isinstance(mkt, dict): mkt = {}
            data.append([str(i), self._Paragraph(_safe_str(p.get("name", p.get("title", "")), 30), self.styles['SmallText']),
                f"{mkt.get('marketing_score', 0):.0%}", str(len(mkt.get("problems", []))),
                str(len(mkt.get("solutions", []))), str(len(mkt.get("recommended_strategies", []))),
                self._Paragraph(_safe_str(mkt.get("summary", "N/A"), 50), self.styles['SmallText'])])
        table = self._make_table(data, [0.3*self.inch, 1.8*self.inch, 0.6*self.inch, 0.7*self.inch, 0.7*self.inch, 0.8*self.inch, 2.0*self.inch], header_color='#FF0000')
        elements.append(table)
        return elements

    def _create_supplier_section(self, products):
        elements = []
        elements.append(self._Paragraph("Supplier Sourcing", self.styles['SectionHeader']))
        data = [["#", "Product", "Supplier", "Price", "MOQ", "Lead Time", "Rating", "Source"]]
        top = sorted(products, key=lambda x: x.get("ai_score", 0), reverse=True)[:20]
        for i, p in enumerate(top, 1):
            data.append([str(i), self._Paragraph(_safe_str(p.get("name", p.get("title", "")), 30), self.styles['SmallText']),
                _safe_str(p.get("supplier_name", "N/A"), 20), _fmt_currency(p.get("supplier_price", 0)),
                str(p.get("supplier_moq", "N/A")), _safe_str(p.get("supplier_lead_time", "N/A"), 15),
                str(p.get("supplier_rating", "N/A")), p.get("supplier_price_source", "N/A")])
        table = self._make_table(data, [0.3*self.inch, 1.8*self.inch, 1.2*self.inch, 0.7*self.inch, 0.6*self.inch, 0.9*self.inch, 0.5*self.inch, 0.8*self.inch], header_color='#548235')
        elements.append(table)
        return elements

    def _create_category_section(self, products):
        elements = []
        elements.append(self._Paragraph("Category Analysis", self.styles['SectionHeader']))
        categories = defaultdict(list)
        for p in products:
            categories[p.get("category", "Unknown")].append(p)
        data = [["Category", "Products", "Avg Price", "Avg Margin", "Avg AI", "Green", "Viable"]]
        for cat, items in sorted(categories.items(), key=lambda x: -len(x[1])):
            avg_price = sum(i.get("amazon_price", 0) for i in items) / max(len(items), 1)
            avg_margin = sum(i.get("estimated_margin_pct", 0) for i in items) / max(len(items), 1)
            avg_ai = sum(i.get("ai_score", 0) for i in items) / max(len(items), 1)
            green_c = sum(1 for i in items if i.get("traffic_light") == "GREEN")
            viable_c = sum(1 for i in items if i.get("viable"))
            data.append([cat[:15], str(len(items)), f"£{avg_price:.2f}", f"{avg_margin:.0f}%", f"{avg_ai:.0%}", str(green_c), str(viable_c)])
        table = self._make_table(data, [1.2*self.inch, 0.8*self.inch, 0.9*self.inch, 0.9*self.inch, 0.7*self.inch, 0.6*self.inch, 0.7*self.inch], header_color='#548235')
        elements.append(table)
        return elements

    def _create_hidden_gems_section(self, gems):
        elements = []
        elements.append(self._Paragraph("Hidden Gems & Emerging Niches", self.styles['SectionHeader']))
        sorted_gems = sorted(gems, key=lambda x: x.get("potential_score", 0), reverse=True)
        data = [["#", "Name", "Potential", "Trend", "Social", "Niche", "Type", "Reasons"]]
        for i, gem in enumerate(sorted_gems[:20], 1):
            reasons = gem.get("reasons", [])
            data.append([str(i), self._Paragraph(_safe_str(gem.get("name", ""), 35), self.styles['SmallText']),
                f"{gem.get('potential_score', 0):.0%}", f"{gem.get('trend_score', 0):.0%}",
                f"{gem.get('social_score', 0):.0%}", f"{gem.get('niche_score', 0):.0%}",
                gem.get("type", ""), self._Paragraph(_safe_str(", ".join(str(r) for r in reasons), 40), self.styles['SmallText'])])
        table = self._make_table(data, [0.3*self.inch, 2.0*self.inch, 0.7*self.inch, 0.6*self.inch, 0.6*self.inch, 0.6*self.inch, 0.9*self.inch, 2.0*self.inch], header_color='#BF8F00')
        elements.append(table)
        return elements
