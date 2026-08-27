"""Master Excel export - all data in a single workbook."""

import os
from datetime import datetime
from typing import Any, Dict, List, Optional

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


PRIORITY_COLORS = {
    "CRITICAL": "FF0000",
    "HIGH": "FFA500",
    "MEDIUM": "FFFF00",
    "LOW": "0000FF",
    "MINIMAL": "808080",
}

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def export_master_excel(
    ideas: List[Dict[str, Any]],
    hidden_gems: Optional[List[Dict[str, Any]]] = None,
    suppliers: Optional[List[Dict[str, Any]]] = None,
    supplier_products: Optional[List[Dict[str, Any]]] = None,
    pricing: Optional[List[Dict[str, Any]]] = None,
    charts: Optional[Dict[str, str]] = None,
    output_path: str = "",
    filename: str = "amazon_product_ideas.xlsx",
) -> str:
    """Export everything to a single Excel workbook."""
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl not installed. Run: pip install openpyxl")

    if hidden_gems is None:
        hidden_gems = []
    if suppliers is None:
        suppliers = []
    if supplier_products is None:
        supplier_products = []
    if pricing is None:
        pricing = []
    if charts is None:
        charts = {}

    wb = Workbook()

    _create_dashboard_sheet(wb, ideas, hidden_gems, suppliers)
    _create_products_sheet(wb, ideas)
    _create_hidden_gems_sheet(wb, hidden_gems)
    _create_marketing_sheet(wb, ideas)
    _create_suppliers_sheet(wb, suppliers)
    _create_supplier_products_sheet(wb, supplier_products)
    _create_profit_calculator_sheet(wb, ideas, pricing)
    _create_pricing_strategy_sheet(wb, ideas, pricing)
    _create_seo_sheet(wb, ideas)
    _create_raw_data_sheet(wb, ideas)

    if charts:
        _create_charts_sheet(wb, charts)

    os.makedirs(output_path, exist_ok=True)
    filepath = os.path.join(output_path, filename)
    wb.save(filepath)
    return filepath


def _create_dashboard_sheet(wb, ideas, hidden_gems, suppliers):
    """Create summary dashboard sheet."""
    ws = wb.active
    ws.title = "Dashboard"

    ws.merge_cells("A1:H1")
    title = ws["A1"]
    title.value = "Amazon Product Idea Generator - Dashboard"
    title.font = Font(bold=True, size=16, color="2F5496")
    title.alignment = Alignment(horizontal="center")

    ws["A3"] = "Generated:"
    ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws["A3"].font = Font(bold=True)

    ws["A5"] = "SUMMARY"
    ws["A5"].font = Font(bold=True, size=14)

    summary_data = [
        ("Total Products Analyzed", len(ideas)),
        ("Hidden Gems Found", len(hidden_gems)),
        ("Suppliers in Database", len(suppliers)),
    ]

    critical = sum(1 for i in ideas if i.get("priority", {}).get("tier") == "CRITICAL")
    high = sum(1 for i in ideas if i.get("priority", {}).get("tier") == "HIGH")
    medium = sum(1 for i in ideas if i.get("priority", {}).get("tier") == "MEDIUM")

    summary_data.extend([
        ("Critical Priority", critical),
        ("High Priority", high),
        ("Medium Priority", medium),
    ])

    if ideas:
        avg_margin = sum(i.get("estimated_margin_pct", 0) for i in ideas) / len(ideas)
        avg_score = sum(i.get("score", 0) for i in ideas) / len(ideas)
        summary_data.extend([
            ("Average Margin", f"{avg_margin:.1f}%"),
            ("Average Score", f"{avg_score:.2f}"),
        ])

    for i, (label, value) in enumerate(summary_data):
        row = 7 + i
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)

    ws["A20"] = "TOP 10 PRODUCTS"
    ws["A20"].font = Font(bold=True, size=14)

    headers = ["Rank", "Name", "Price", "Margin", "Score", "Priority", "Reviews"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=22, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    for i, idea in enumerate(ideas[:10], 1):
        row = 22 + i
        priority = idea.get("priority", {})
        data = [
            priority.get("rank", i),
            idea.get("name", "")[:50],
            idea.get("amazon_price", 0) or idea.get("price", 0),
            idea.get("estimated_margin_pct", 0),
            idea.get("score", 0),
            priority.get("tier", ""),
            idea.get("review_count", 0),
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = THIN_BORDER
            if col in [3, 4, 5]:
                cell.number_format = "0.00"

    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 20


def _create_products_sheet(wb, ideas):
    """Create detailed products sheet."""
    ws = wb.create_sheet("Product Ideas")

    headers = [
        "Priority Rank", "Priority Tier", "Product Name", "ASIN", "Category",
        "Price (£)", "Rating", "Reviews", "Margin (%)", "Access",
        "Score", "Marketing Score", "Action", "URL", "Image URL"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER

    for row_idx, idea in enumerate(ideas, 2):
        priority = idea.get("priority", {})
        marketing = idea.get("marketing", {})
        tier = priority.get("tier", "unknown")

        data = [
            priority.get("rank", row_idx - 1),
            tier,
            idea.get("name", ""),
            idea.get("asin", ""),
            idea.get("category", ""),
            idea.get("amazon_price", 0) or idea.get("price", 0),
            idea.get("rating", 0),
            idea.get("review_count", 0),
            idea.get("estimated_margin_pct", 0),
            "Gated" if idea.get("gated") else "Ungated",
            idea.get("score", 0),
            marketing.get("marketing_score", 0),
            priority.get("action", ""),
            idea.get("url", ""),
            idea.get("image", ""),
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = THIN_BORDER
            if col in [6, 7, 8, 9, 11, 12]:
                cell.number_format = "0.00"
            if col == 15 and value:
                cell.hyperlink = value
                cell.font = Font(color="0563C1", underline="single")

        if tier in PRIORITY_COLORS:
            tier_cell = ws.cell(row=row_idx, column=2)
            tier_cell.fill = PatternFill(
                start_color=PRIORITY_COLORS[tier],
                end_color=PRIORITY_COLORS[tier],
                fill_type="solid"
            )
            if tier in ["CRITICAL", "HIGH"]:
                tier_cell.font = Font(bold=True, color="FFFFFF")

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(ideas) + 1}"
    ws.freeze_panes = "A2"

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18


def _create_hidden_gems_sheet(wb, hidden_gems):
    """Create hidden gems sheet."""
    ws = wb.create_sheet("Hidden Gems")

    headers = [
        "Potential Score", "Product Name", "ASIN", "Price",
        "Reviews", "Type", "Trend Score", "Social Score",
        "Niche Score", "Margin Score", "Reasons"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
        cell.border = THIN_BORDER

    for row_idx, gem in enumerate(hidden_gems, 2):
        data = [
            gem.get("potential_score", 0),
            gem.get("name", ""),
            gem.get("asin", ""),
            gem.get("amazon_price", 0),
            gem.get("review_count", 0),
            gem.get("type", ""),
            gem.get("trend_score", 0),
            gem.get("social_score", 0),
            gem.get("niche_score", 0),
            gem.get("margin_score", 0),
            ", ".join(gem.get("reasons", [])),
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = THIN_BORDER
            if col in [1, 4, 7, 8, 9, 10]:
                cell.number_format = "0.000"

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(hidden_gems) + 1}"
    ws.freeze_panes = "A2"

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18


def _create_marketing_sheet(wb, ideas):
    """Create marketing analysis sheet."""
    ws = wb.create_sheet("Marketing Analysis")

    headers = [
        "Product Name", "Marketing Score", "Summary",
        "Problems", "Solutions", "Recommended Strategies"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
        cell.border = THIN_BORDER

    for row_idx, idea in enumerate(ideas, 2):
        marketing = idea.get("marketing", {})
        problems = marketing.get("problems", [])
        solutions = marketing.get("solutions", [])
        strategies = marketing.get("recommended_strategies", [])

        problems_text = "\n".join([
            f"- [{p.get('severity', '').upper()}] {p.get('problem', '')}"
            for p in problems
        ])

        solutions_text = "\n".join([
            f"- {s.get('solution', '')} ({s.get('priority', '')})"
            for s in solutions
        ])

        strategies_text = "\n".join([
            f"- {st.get('name', '')} ({st.get('priority', '')})"
            for st in strategies
        ])

        data = [
            idea.get("name", ""),
            marketing.get("marketing_score", 0),
            marketing.get("summary", ""),
            problems_text,
            solutions_text,
            strategies_text,
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = [30, 15, 40, 40, 40, 40]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def _create_suppliers_sheet(wb, suppliers):
    """Create suppliers sheet."""
    ws = wb.create_sheet("Suppliers")

    headers = [
        "ID", "Name", "Company", "Country", "Location", "Website",
        "Contact Person", "Email", "Phone", "WhatsApp", "Skype", "WeChat",
        "Business Type", "Year Established", "Employees", "Certifications",
        "MOQ", "Lead Time (Days)", "Payment Terms", "Shipping Methods",
        "Rating", "Notes"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
        cell.border = THIN_BORDER

    for row_idx, supplier in enumerate(suppliers, 2):
        data = [
            supplier.get("id", ""),
            supplier.get("name", ""),
            supplier.get("company_name", ""),
            supplier.get("country", ""),
            supplier.get("location", ""),
            supplier.get("website", ""),
            supplier.get("contact_person", ""),
            supplier.get("contact_email", ""),
            supplier.get("contact_phone", ""),
            supplier.get("contact_whatsapp", ""),
            supplier.get("contact_skype", ""),
            supplier.get("contact_wechat", ""),
            supplier.get("business_type", ""),
            supplier.get("year_established", ""),
            supplier.get("employee_count", ""),
            supplier.get("certifications", ""),
            supplier.get("moq", ""),
            supplier.get("lead_time_days", ""),
            supplier.get("payment_terms", ""),
            supplier.get("shipping_methods", ""),
            supplier.get("rating", 0),
            supplier.get("notes", ""),
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = THIN_BORDER

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(suppliers) + 1}"
    ws.freeze_panes = "A2"


def _create_supplier_products_sheet(wb, supplier_products):
    """Create supplier products sheet."""
    ws = wb.create_sheet("Supplier Products")

    headers = [
        "Supplier", "Product Name", "ASIN", "SKU", "Unit Cost",
        "Shipping Cost", "Min Order", "Bulk Prices", "Notes"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
        cell.border = THIN_BORDER

    for row_idx, product in enumerate(supplier_products, 2):
        bulk = product.get("bulk_prices", {})
        if isinstance(bulk, dict):
            bulk_str = ", ".join([f"{k}: £{v}" for k, v in bulk.items()])
        else:
            bulk_str = str(bulk)

        data = [
            product.get("supplier_name", ""),
            product.get("product_name", ""),
            product.get("asin", ""),
            product.get("sku", ""),
            product.get("unit_cost", 0),
            product.get("shipping_cost", 0),
            product.get("min_order", 1),
            bulk_str,
            product.get("notes", ""),
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = THIN_BORDER
            if col in [5, 6]:
                cell.number_format = "£#,##0.00"

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(supplier_products) + 1}"


def _create_profit_calculator_sheet(wb, ideas, pricing):
    """Create profit calculator sheet."""
    ws = wb.create_sheet("Profit Calculator")

    headers = [
        "Product Name", "ASIN", "Supplier", "Supplier Cost", "Shipping",
        "Customs", "FBA Fee", "Referral Fee", "Total Landed",
        "Selling Price", "Profit/Unit", "Margin %", "ROI %", "Break-even"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        cell.border = THIN_BORDER

    pricing_map = {p.get("asin"): p for p in pricing}

    for row_idx, idea in enumerate(ideas, 2):
        asin = idea.get("asin", "")
        p = pricing_map.get(asin, {})

        data = [
            idea.get("name", ""),
            asin,
            p.get("supplier_name", ""),
            p.get("supplier_cost", 0),
            p.get("shipping_cost", 0),
            p.get("customs_duty", 0),
            p.get("fba_fee", 0),
            p.get("referral_fee", 0),
            p.get("total_landed_cost", 0),
            idea.get("amazon_price", 0) or idea.get("price", 0),
            p.get("profit_per_unit", 0),
            p.get("margin_percent", 0),
            p.get("roi_percent", 0),
            p.get("break_even_units", 0),
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = THIN_BORDER
            if col in [4, 5, 6, 7, 8, 9, 10, 11, 12, 13]:
                cell.number_format = "£#,##0.00" if col != 12 and col != 13 else "0.00%"

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(ideas) + 1}"
    ws.freeze_panes = "A2"


def _create_pricing_strategy_sheet(wb, ideas, pricing):
    """Create pricing strategy sheet."""
    ws = wb.create_sheet("Pricing Strategy")

    headers = [
        "Product Name", "ASIN", "Market Price", "Min Price",
        "Suggested Price", "Optimal Price", "Max Price",
        "Target Margin", "Actual Margin"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        cell.border = THIN_BORDER

    pricing_map = {p.get("asin"): p for p in pricing}

    for row_idx, idea in enumerate(ideas, 2):
        asin = idea.get("asin", "")
        p = pricing_map.get(asin, {})
        market_price = idea.get("amazon_price", 0) or idea.get("price", 0)

        data = [
            idea.get("name", ""),
            asin,
            market_price,
            p.get("min_price", market_price * 0.7),
            p.get("suggested_price", market_price * 0.95),
            p.get("suggested_price", market_price * 0.95),
            p.get("max_price", market_price * 1.2),
            p.get("target_margin", 30),
            idea.get("estimated_margin_pct", 0),
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = THIN_BORDER
            if col in [3, 4, 5, 6, 7, 8, 9]:
                cell.number_format = "£#,##0.00" if col < 8 else "0.00%"

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(ideas) + 1}"


def _create_raw_data_sheet(wb, ideas):
    """Create raw data sheet with all collected information."""
    ws = wb.create_sheet("Raw Data")

    headers = [
        "ASIN", "Name", "Category", "Price", "Rating", "Reviews",
        "URL", "Image", "Source", "Query", "Score"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
        cell.border = THIN_BORDER

    for row_idx, idea in enumerate(ideas, 2):
        data = [
            idea.get("asin", ""),
            idea.get("name", ""),
            idea.get("category", ""),
            idea.get("amazon_price", 0) or idea.get("price", 0),
            idea.get("rating", 0),
            idea.get("review_count", 0),
            idea.get("url", ""),
            idea.get("image", ""),
            idea.get("source", ""),
            idea.get("query", ""),
            idea.get("score", 0),
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = THIN_BORDER

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(ideas) + 1}"


def _create_seo_sheet(wb, ideas):
    """Create SEO keywords sheet."""
    ws = wb.create_sheet("SEO Keywords")

    headers = [
        "Product Name", "ASIN", "SEO Score", "Optimized Title",
        "Primary Keywords", "Long-tail Keywords", "Backend Keywords",
        "Search Terms", "Bullet Keywords"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
        cell.border = THIN_BORDER

    for row_idx, idea in enumerate(ideas, 2):
        listing = idea.get("listing_template", {})
        seo = listing.get("seo", {})
        identity = listing.get("product_identity", {})

        primary = ", ".join(seo.get("primary_keywords", [])[:8])
        long_tail = ", ".join(seo.get("long_tail_keywords", [])[:5])
        backend = ", ".join(seo.get("backend_keywords", [])[:8])
        search_terms = seo.get("search_terms", "")
        seo_score = seo.get("seo_score", {}).get("percentage", 0)

        bullets = listing.get("description", {}).get("bullet_points", [])
        bullet_text = " | ".join([b[:50] for b in bullets if b][:5])

        data = [
            idea.get("name", ""),
            idea.get("asin", ""),
            f"{seo_score:.0f}%",
            identity.get("item_name", ""),
            primary,
            long_tail,
            backend,
            search_terms,
            bullet_text,
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(ideas) + 1}"
    ws.freeze_panes = "A2"

    widths = [30, 15, 10, 40, 40, 40, 40, 40, 40]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def _create_charts_sheet(wb, charts):
    """Create charts sheet with embedded chart images."""
    ws = wb.create_sheet("Charts")

    ws["A1"] = "Charts Export"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A3"] = "Chart files have been saved to:"
    ws["A4"] = os.path.join(os.path.dirname(__file__), "..", "output", "charts")

    row = 6
    for name, path in charts.items():
        ws.cell(row=row, column=1, value=name).font = Font(bold=True)
        ws.cell(row=row, column=2, value=path)
        row += 1
